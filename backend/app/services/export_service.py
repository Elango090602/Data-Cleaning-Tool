import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def export_csv(df: pd.DataFrame, path: str):
    """
    Saves a Pandas DataFrame to a CSV file. Uses utf-8-sig for Excel compatibility.
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    df.to_csv(path, index=False, encoding='utf-8-sig')

def export_xlsx(df: pd.DataFrame, path: str):
    """
    Saves a Pandas DataFrame to an Excel file.
    - Freezes the top row header.
    - Autoscale column widths to match contents.
    - Styles cells professionally (Segoe UI, subtle colors).
    """
    os.makedirs(os.path.dirname(path), exist_ok=True)
    
    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Leads')
        workbook = writer.book
        worksheet = writer.sheets['Leads']
        
        # Freeze the top header row
        worksheet.freeze_panes = 'A2'
        
        # Aesthetics: styling headers
        header_fill = PatternFill(start_color='ECEEF0', end_color='ECEEF0', fill_type='solid') # surface-container
        header_font = Font(name='Segoe UI', size=11, bold=True, color='191C1E')
        header_align = Alignment(horizontal='left', vertical='center')
        
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            
        # Styling data cells slightly
        data_font = Font(name='Segoe UI', size=10, color='191C1E')
        data_align = Alignment(horizontal='left', vertical='center')
        
        for row in range(2, worksheet.max_row + 1):
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = data_align
                
        # Auto-fit columns
        for col in worksheet.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            # Set appropriate column width
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 10)

def generate_summary_report(summary_dict: dict, path: str, format_type: str = 'csv'):
    """
    Generates a clean tabular summary file from summary dictionary metrics.
    """
    df = pd.DataFrame([
        {"Metric": "Total Uploaded Records", "Value": summary_dict.get("total_uploaded", 0)},
        {"Metric": "Total Records After Cleaning", "Value": summary_dict.get("total_after_cleaning", 0)},
        {"Metric": "Valid Records", "Value": summary_dict.get("valid_records", 0)},
        {"Metric": "Records Needing Review", "Value": summary_dict.get("needs_review", 0)},
        {"Metric": "Invalid Records", "Value": summary_dict.get("invalid_records", 0)},
        {"Metric": "Duplicates Found", "Value": summary_dict.get("duplicates_found", 0)},
        {"Metric": "Duplicates Removed", "Value": summary_dict.get("duplicates_removed", 0)},
        {"Metric": "Invalid Emails", "Value": summary_dict.get("invalid_emails", 0)},
        {"Metric": "Invalid Phones", "Value": summary_dict.get("invalid_phones", 0)},
        {"Metric": "Processing Time (ms)", "Value": summary_dict.get("processing_time_ms", 0)}
    ])
    
    if format_type == 'xlsx':
        export_xlsx(df, path)
    else:
        export_csv(df, path)
